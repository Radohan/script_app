import os
import sys
import re
import traceback
import codecs
import webbrowser
from PyQt5.QtWidgets import (QMainWindow, QVBoxLayout, QWidget, QFileDialog,
                           QTableWidgetItem, QMessageBox, QStatusBar, QHeaderView,
                           QPushButton, QAction, QFrame, QHBoxLayout, QToolBar,
                           QMenu, QToolButton, QSizePolicy, QProgressBar, QLabel,
                           QApplication)
from PyQt5.QtCore import Qt, QTimer, QSize, QPoint
from PyQt5.QtGui import QFont, QColor, QPainter

from ui.theme import ThemeManager
from ui.ui_components import UIComponents
from ui.custom_widgets import TranslationDiffDialog
from utils.utils import (extract_main_key, extract_line_number, has_comments,
                         get_comment_text, natural_sort_key, find_text_differences)
from utils.xml_parser import XMLParser
from utils.document_parser import DocumentParser
from utils.FileProcessingWorker import FileProcessingWorker
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import pandas

class MXLIFFParser(QMainWindow):
    def __init__(self):
        super().__init__()

        # Load custom fonts
        self.load_fonts()

        # Dark mode state
        self.dark_mode = False

        # Diff highlighting state
        self.diff_highlighting_enabled = True

        # Define fixed column configuration in the exact order shown
        self.default_columns = [
            'Key',
            'Info',
            'Speaker',
            'Source Text',
            'Target Text',
            'Char Info',  # New column for character count and comment icons
            'Speaker and Target',
            'Player Info'
        ]

        # Current column order is now fixed
        self.current_columns = self.default_columns.copy()

        # Column indices map (maps logical columns to visual positions)
        self.column_map = {col: idx for idx, col in enumerate(self.current_columns)}

        # Define color themes
        self.light_theme = ThemeManager.get_light_theme()
        self.dark_theme = ThemeManager.get_dark_theme()

        # Set current theme
        self.current_theme = self.light_theme

        # Add these variables for caching
        self.original_xml_content = None  # Store the original XML content
        self.current_file_path = None  # Store the current file path

        # Initialize variables
        self.group_headers = []
        self.group_rows = {}
        self.processed_data = []  # Store the processed data for reuse
        self.diff_pairs = {}  # Store pairs of related translations
        self.updating_cell = False  # Flag to prevent recursive editing

        self.initUI()

    # Add these methods to your MXLIFFParser class

    def open_file(self):
        """Open an MXLIFF file and parse it."""
        # Check for unsaved changes
        if self.has_unsaved_changes():
            reply = QMessageBox.question(
                self,
                "Unsaved Changes",
                "You have unsaved changes. Do you want to save them before opening a new file?",
                QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel,
                QMessageBox.Save
            )

            if reply == QMessageBox.Save:
                self.export_file()
                # If user cancels during export, abort opening
                if self.has_unsaved_changes():
                    return
            elif reply == QMessageBox.Cancel:
                return

        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Open MXLIFF File",
            "",
            "MXLIFF Files (*.mxliff *.xml);;All Files (*)",
            options=options
        )

        if file_path:
            self.log(f"Opening file: {file_path}")
            self.file_label.setText(os.path.basename(file_path))

            # Start processing in a worker thread
            self._start_xml_parsing(file_path)

    def _start_xml_parsing(self, file_path):
        """Start XML parsing in a worker thread."""
        # Show progress bar
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.progress_bar.setRange(0, 100)  # Set to determinate mode
        self.statusBar.showMessage("Processing file...")

        # Clear previous results
        self.table.setRowCount(0)
        self.processed_data = []
        self.diff_pairs = {}

        # Create worker thread
        self.worker = FileProcessingWorker(file_path, 'parse_xml', self)

        # Connect signals
        self.worker.progress_signal.connect(self._update_progress)
        self.worker.finished_signal.connect(self._on_xml_parsed)
        self.worker.error_signal.connect(self._on_worker_error)

        # Start worker
        self.worker.start()

    def create_info_icon(self, row, column, tooltip=""):
        """Create an info icon in the specified cell with optional tooltip."""
        from PyQt5.QtGui import QIcon, QPixmap
        from PyQt5.QtCore import QSize, Qt

        # Create a label widget to hold the icon
        from PyQt5.QtWidgets import QLabel

        icon_label = QLabel()
        icon_label.setAlignment(Qt.AlignCenter)

        # Create the icon programmatically (since we can't include external images)
        pixmap = QPixmap(24, 24)
        pixmap.fill(Qt.transparent)

        # Draw the info icon
        painter = QPainter(pixmap)
        painter.setRenderHint(QPainter.Antialiasing)

        # Draw circle
        painter.setPen(Qt.black)
        painter.setBrush(Qt.white)
        painter.drawEllipse(2, 2, 20, 20)

        # Draw 'i' letter
        font = QFont("Arial", 14, QFont.Bold)
        painter.setFont(font)
        painter.drawText(pixmap.rect(), Qt.AlignCenter, "i")
        painter.end()

        # Set the icon
        icon_label.setPixmap(pixmap)

        # Set tooltip if provided
        if tooltip:
            icon_label.setToolTip(tooltip)

        # Add the label to the table
        self.table.setCellWidget(row, column, icon_label)

        return icon_label

    def _update_progress(self, value, message):
        """Update progress bar and status message."""
        self.progress_bar.setValue(value)
        self.statusBar.showMessage(message)

        # Process events to keep UI responsive during progress updates
        QApplication.processEvents()

    def _on_xml_parsed(self, result):
        # Code processing the result...

        # Enable export controls now that we have a file loaded
        if hasattr(self, 'export_button'):
            self.export_button.setEnabled(True)
            self.log("MXLIFF export button enabled")
        if hasattr(self, 'export_action'):
            self.export_action.setEnabled(True)

        # Enable Excel export controls - add debug logging
        if hasattr(self, 'export_excel_button'):
            self.export_excel_button.setEnabled(True)
            self.log("Excel export button enabled")
        if hasattr(self, 'export_excel_action'):
            self.export_excel_action.setEnabled(True)
            self.log("Excel export action enabled")

        # Also enable the button in the regular file loading method
        if hasattr(self, '_process_file'):
            # After successful loading
            if hasattr(self, 'export_excel_button'):
                self.export_excel_button.setEnabled(True)
                self.log("Excel export button enabled in _process_file")

    def _on_worker_error(self, error_message):
        """Handle worker thread errors."""
        # Hide progress bar
        self.progress_bar.setVisible(False)

        # Update status
        self.statusBar.showMessage("Error processing file", 5000)
        self.file_label.setText(f"Error loading file")

        # Log error
        self.log(f"Worker error: {error_message}")

        # Show error message
        QMessageBox.critical(
            self,
            "Error Processing File",
            f"An error occurred while processing the file:\n\n{error_message}"
        )

    def update_comments_in_display(self, updated_keys):
        """Update only the items with new comments without rebuilding the entire table."""
        if not updated_keys:
            return  # Nothing to update

        # Find the rows that need updating
        key_col = self.column_map.get('Key', 0)
        source_col = self.column_map.get('Source Text', 2)

        for row in range(self.table.rowCount()):
            key_item = self.table.item(row, key_col)
            if not key_item:
                continue

            key_text = key_item.text()
            if key_text in updated_keys:
                # Find the corresponding data item
                data_item = None
                for data in self.processed_data:
                    if not data.get('is_header', True) and 'item' in data:
                        item = data['item']
                        if item.get('key', '') == key_text:
                            data_item = item
                            break

                if data_item:
                    # Update tooltip with new note text
                    has_comment = has_comments(data_item)
                    tooltip_text = data_item.get('note_text', '')
                    if has_comment:
                        comment_text = get_comment_text(data_item)
                        tooltip_text = f"{comment_text}\n\n{tooltip_text}"

                    # Update all cells in this row with the new tooltip
                    for col in range(self.table.columnCount()):
                        cell = self.table.item(row, col)
                        if cell:
                            cell.setToolTip(tooltip_text)

                    # Add a visual indicator that this item has comments in the Source Text column
                    if source_col >= 0:
                        source_item = self.table.item(row, source_col)
                        if source_item:
                            current_text = source_item.text()
                            if has_comment and not current_text.startswith('💬 '):
                                source_item.setText("💬 " + current_text.replace('💬 ', ''))
                            elif not has_comment and current_text.startswith('💬 '):
                                # Remove icon if comment was removed
                                source_item.setText(current_text.replace('💬 ', ''))

    def export_file(self):
        """Export the updated MXLIFF file with edited translations."""
        if not self.original_xml_content or not self.current_file_path:
            QMessageBox.warning(
                self,
                "Export Error",
                "No file is currently loaded or the original content is missing."
            )
            return

        # Check if we have any changes to export
        has_changes = False
        for data in self.processed_data:
            if not data.get('is_header', True) and 'item' in data:
                item = data['item']
                if 'target_text' in item and 'original_target_text' in item:
                    if item['target_text'] != item['original_target_text']:
                        has_changes = True
                        break

        if not has_changes:
            reply = QMessageBox.question(
                self,
                "No Changes Detected",
                "No changes to translations were detected. Do you still want to export the file?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply == QMessageBox.No:
                return

        # Choose where to save the file
        options = QFileDialog.Options()
        default_name = "edited_" + os.path.basename(self.current_file_path)
        save_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export MXLIFF File",
            default_name,
            "MXLIFF Files (*.mxliff *.xml);;All Files (*)",
            options=options
        )

        if not save_path:
            return  # User cancelled

        # Start export in worker thread
        self._start_export(save_path)

    def _start_export(self, save_path):
        """Start export in a worker thread."""
        # Show progress bar
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.progress_bar.setRange(0, 100)  # Set to determinate mode
        self.statusBar.showMessage("Exporting file...")

        # Create worker thread
        self.worker = FileProcessingWorker(save_path, 'export_file', self)

        # Set data needed for export
        self.worker.set_data('processed_data', self.processed_data)
        self.worker.set_data('original_xml_content', self.original_xml_content)

        # Connect signals
        self.worker.progress_signal.connect(self._update_progress)
        self.worker.finished_signal.connect(self._on_export_completed)
        self.worker.error_signal.connect(self._on_worker_error)

        # Start worker
        self.worker.start()

    def export_to_excel(self):
        """Export the current table data to Excel with formatting preserved."""
        try:
            self.log("Export to Excel button clicked!")

            # Force the button to show as enabled in case it's not
            if hasattr(self, 'export_excel_button'):
                self.export_excel_button.setEnabled(True)
                self.export_excel_button.repaint()  # Force UI update

            if not hasattr(self, 'processed_data') or not self.processed_data:
                self.log("No data to export")
                QMessageBox.warning(
                    self,
                    "No Data to Export",
                    "Please open an MXLIFF file first before exporting to Excel."
                )
                return

            # Choose where to save the Excel file
            self.log("Opening file dialog")
            options = QFileDialog.Options()
            default_name = os.path.splitext(os.path.basename(self.current_file_path))[0] + ".xlsx"
            save_path, _ = QFileDialog.getSaveFileName(
                self,
                "Export to Excel",
                default_name,
                "Excel Files (*.xlsx);;All Files (*)",
                options=options
            )

            if not save_path:
                self.log("User cancelled file dialog")
                return  # User cancelled

            # Show progress
            self.log(f"Starting export to {save_path}")
            self.progress_bar.setVisible(True)
            self.statusBar.showMessage("Exporting to Excel...")

            # We'll export in a timer to allow UI to update
            self.log("Setting up timer for _process_excel_export")
            QTimer.singleShot(100, lambda: self._process_excel_export(save_path))

        except Exception as e:
            self.log(f"Error in export_to_excel: {str(e)}")
            self.log(traceback.format_exc())
            self.progress_bar.setVisible(False)
            error_msg = f"Error preparing Excel export: {str(e)}"
            self.statusBar.showMessage(error_msg, 5000)

            QMessageBox.critical(
                self,
                "Export Error",
                error_msg
            )

    def _process_excel_export(self, save_path):
        """Process the Excel export after a short delay to allow UI to update."""
        try:
            # Create a new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "MXLIFF Data"

            # Add an additional column for notes/comments
            notes_column_name = "Additional Notes"

            # Define column headers based on the current column order
            headers = self.current_columns.copy()
            headers.append(notes_column_name)  # Add notes column

            for i, header in enumerate(headers, 1):
                ws.cell(row=1, column=i, value=header)
                ws.cell(row=1, column=i).font = Font(bold=True)
                # Add a light gray background to header row
                ws.cell(row=1, column=i).fill = PatternFill(start_color="D3D3D3", fill_type="solid")

            # Get theme colors for formatting
            group_header_color = self._excel_color_from_qcolor(QColor(self.current_theme['group_header']))
            menu_label_color = self._excel_color_from_qcolor(QColor(self.current_theme['menu_label']))
            female_key_color = self._excel_color_from_qcolor(QColor(self.current_theme['female_key']))
            diff_text_color = self._excel_color_from_qcolor(QColor(self.current_theme.get('diff_text', '#FF0000')))

            # Track current Excel row
            excel_row = 2  # Start after the header row

            # Process each row in the table and add to Excel
            for row in range(self.table.rowCount()):
                # Check if this is a header row
                first_cell = self.table.item(row, 0)
                if not first_cell:
                    continue  # Skip empty rows

                # Check if this is a group header
                is_header = False
                header_data = first_cell.data(Qt.UserRole) if first_cell else None
                if header_data and isinstance(header_data, dict) and header_data.get('is_header'):
                    is_header = True
                    # Add group header as a merged cell
                    header_text = first_cell.text()
                    ws.cell(row=excel_row, column=1, value=header_text)
                    # Apply header formatting
                    header_cell = ws.cell(row=excel_row, column=1)
                    header_cell.font = Font(bold=True)
                    header_cell.fill = PatternFill(start_color=group_header_color,
                                                   fill_type="solid")
                    # Merge cells across all columns
                    ws.merge_cells(start_row=excel_row, start_column=1, end_row=excel_row, end_column=len(headers))
                    excel_row += 1
                    continue

                # Check if this is a scene info row
                if first_cell and 'Scene:' in first_cell.text() and self.table.columnSpan(row, 0) > 1:
                    scene_text = first_cell.text()
                    ws.cell(row=excel_row, column=1, value=scene_text)
                    # Apply scene info formatting
                    scene_cell = ws.cell(row=excel_row, column=1)
                    scene_cell.font = Font(italic=True)
                    lighter_color = self._lighten_excel_color(group_header_color)
                    scene_cell.fill = PatternFill(start_color=lighter_color, fill_type="solid")
                    # Merge cells across all columns
                    ws.merge_cells(start_row=excel_row, start_column=1, end_row=excel_row, end_column=len(headers))
                    excel_row += 1
                    continue

                # Check if this is a missing line row
                if first_cell and '[MISSING LINE' in first_cell.text() and self.table.columnSpan(row, 0) > 1:
                    missing_text = first_cell.text()
                    ws.cell(row=excel_row, column=1, value=missing_text)
                    # Apply missing line formatting
                    missing_cell = ws.cell(row=excel_row, column=1)
                    missing_cell.font = Font(bold=True, color="AA0000")
                    missing_cell.fill = PatternFill(start_color="FFCCCC", fill_type="solid")
                    missing_cell.alignment = Alignment(horizontal='center', vertical='center')
                    # Merge cells across all columns
                    ws.merge_cells(start_row=excel_row, start_column=1, end_row=excel_row, end_column=len(headers))
                    excel_row += 1
                    continue

                # Regular data row
                cell_colors = {}

                # Get key to determine row type
                key_col = self.column_map.get('Key', 0)
                key_item = self.table.item(row, key_col)
                key_text = key_item.text() if key_item else ""

                # Determine row background color
                bg_color = None
                if key_item:
                    if 'MenuLabel' in key_text:
                        bg_color = menu_label_color
                    elif key_text.endswith('.F'):
                        bg_color = female_key_color

                # Prepare to collect notes/additional information
                row_notes = []

                # Process each column
                for col_name, col_index in self.column_map.items():
                    cell = self.table.item(row, col_index)
                    if not cell:
                        continue

                    excel_col = self.current_columns.index(col_name) + 1  # Excel is 1-indexed
                    cell_value = cell.text()

                    # Create Excel cell
                    excel_cell = ws.cell(row=excel_row, column=excel_col, value=cell_value)

                    # Apply basic formatting
                    excel_cell.alignment = Alignment(wrap_text=True, vertical='top')

                    # Apply background color if needed
                    if bg_color:
                        excel_cell.fill = PatternFill(start_color=bg_color, fill_type="solid")

                    # Apply text color for diff highlighting
                    if cell.foreground().color().name() == QColor(
                            self.current_theme.get('diff_text', '#FF0000')).name():
                        excel_cell.font = Font(color=diff_text_color)

                    # Bold for edited cells
                    if cell.font().bold():
                        excel_cell.font = Font(bold=True)

                    # Collect tooltip information
                    tooltip = cell.toolTip()
                    if tooltip:
                        row_notes.append(f"{col_name}: {tooltip}")

                # Add collected notes to the last column
                notes_col = len(headers)
                ws.cell(row=excel_row, column=notes_col, value='; '.join(row_notes))

                excel_row += 1

            # Auto-adjust column widths
            for col_idx, col_name in enumerate(headers, 1):
                # Set column width based on content
                if col_name in ['Source Text', 'Target Text', 'Additional Notes']:
                    # Make text columns wider
                    ws.column_dimensions[get_column_letter(col_idx)].width = 60
                elif col_name == 'Char Info':
                    # Make character info column narrower
                    ws.column_dimensions[get_column_letter(col_idx)].width = 15
                else:
                    # Default width for other columns
                    ws.column_dimensions[get_column_letter(col_idx)].width = 25

            # Save the workbook
            wb.save(save_path)

            # Hide progress bar and show success message
            self.progress_bar.setVisible(False)
            self.statusBar.showMessage(f"Excel file exported successfully to {save_path}", 5000)

            # Show success message
            QMessageBox.information(
                self,
                "Export Successful",
                f"Excel file was successfully exported to:\n{save_path}\n\n"
                f"Total rows: {excel_row - 1}\n"
                f"Formatting and additional information have been preserved in the 'Additional Notes' column."
            )

        except Exception as e:
            # Log and show error
            self.log(f"Excel export error: {str(e)}")
            import traceback
            self.log(traceback.format_exc())

            self.progress_bar.setVisible(False)
            self.statusBar.showMessage(f"Error exporting to Excel: {str(e)}", 5000)

            QMessageBox.critical(
                self,
                "Excel Export Error",
                f"An error occurred while exporting to Excel:\n\n{str(e)}"
            )

    def _excel_color_from_qcolor(self, qcolor):
        """Convert a QColor to Excel color string format (RRGGBB)."""
        return f"{qcolor.red():02X}{qcolor.green():02X}{qcolor.blue():02X}"

    def _lighten_excel_color(self, hex_color, factor=0.2):
        """Lighten a hex color by a factor (0.0-1.0)."""
        # Convert hex to RGB
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)

        # Lighten
        r = min(255, int(r + (255 - r) * factor))
        g = min(255, int(g + (255 - g) * factor))
        b = min(255, int(b + (255 - b) * factor))

        # Convert back to hex
        return f"{r:02X}{g:02X}{b:02X}"

    def _on_export_completed(self, result):
        """Handle completed export."""
        # Hide progress bar
        self.progress_bar.setVisible(False)

        # Check for success
        if result.get('success', False):
            file_path = result.get('file_path', '')

            # Update status
            self.statusBar.showMessage(f"File exported successfully to {file_path}", 5000)

            # Calculate editing stats
            edited_count = 0
            for data in self.processed_data:
                if not data.get('is_header', True) and 'item' in data:
                    item = data['item']
                    if 'target_text' in item and 'original_target_text' in item:
                        if item['target_text'] != item['original_target_text']:
                            edited_count += 1

            # Show success message
            QMessageBox.information(
                self,
                "Export Successful",
                f"MXLIFF file was successfully exported to:\n{file_path}\n\n"
                f"Total translations: {len(self.processed_data) - len(self.group_headers)}\n"
                f"Edited translations: {edited_count}"
            )
        else:
            # Show error message
            QMessageBox.critical(
                self,
                "Export Error",
                "Failed to export file. See log for details."
            )

    def upload_document(self):
        """Handle uploading and processing a Word or PDF document."""
        # Check if we have MXLIFF data loaded
        if not hasattr(self, 'processed_data') or not self.processed_data:
            QMessageBox.warning(
                self,
                "No MXLIFF Data",
                "Please open an MXLIFF file first before uploading a document."
            )
            return

        # Initialize document parser for file selection only
        self.document_parser = DocumentParser(self)

        # Select document file
        file_path = self.document_parser.select_document()
        if not file_path:
            return  # User cancelled

        # Always start document processing in worker thread
        self._start_document_processing(file_path)

    def _start_document_processing(self, file_path):
        """Start document processing in a worker thread."""
        # Show progress bar
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.progress_bar.setRange(0, 100)  # Set to determinate mode
        self.statusBar.showMessage(f"Processing document: {os.path.basename(file_path)}...")

        # Add cancel button during processing
        if not hasattr(self, 'cancel_button'):
            self.cancel_button = QPushButton("Cancel")
            self.cancel_button.clicked.connect(self.cancel_current_operation)
            self.statusBar.addPermanentWidget(self.cancel_button)
        else:
            self.cancel_button.show()

        # Create worker thread
        self.worker = FileProcessingWorker(file_path, 'process_document', self)

        # Set data needed for processing
        self.worker.set_data('processed_data', self.processed_data)

        # Connect signals
        self.worker.progress_signal.connect(self._update_progress)
        self.worker.finished_signal.connect(self._on_document_processed)
        self.worker.error_signal.connect(self._on_worker_error)

        # Start worker
        self.worker.start()

    def _on_document_processed(self, result):
        """Handle completed document processing."""
        # Hide progress bar
        self.progress_bar.setVisible(False)

        # Check for success
        if result.get('success', False):
            match_results = result.get('match_results', {'matches': 0, 'updates': []})

            # Process updates
            comment_updates = 0
            updated_keys = set()  # Just for tracking, not for UI updates

            for update in match_results['updates']:
                key = update.get('key')
                comment = update.get('comment')

                # Find the item in processed data
                for data in self.processed_data:
                    if not data.get('is_header', True) and 'item' in data:
                        item = data['item']
                        if item.get('key') == key:
                            # Add comment to note_text field
                            note_text = item.get('note_text', '')

                            # Instead of replacing existing comments, append the new one with "CoT Comment:" prefix
                            if note_text:
                                # Add new comment as a new line
                                note_text += f"\nCoT Comment: {comment}"
                            else:
                                # First comment for this item
                                note_text = f"CoT Comment: {comment}"

                            # Update the note_text
                            item['note_text'] = note_text
                            comment_updates += 1
                            updated_keys.add(key)
                            break

            # Update the Source Text column with comment icons
            key_col = self.column_map.get('Key', 0)
            source_col = self.column_map.get('Source Text', 2)

            for row in range(self.table.rowCount()):
                key_item = self.table.item(row, key_col)
                if not key_item:
                    continue

                key_text = key_item.text()

                # Find if this key has comments
                for data in self.processed_data:
                    if not data.get('is_header', True) and 'item' in data:
                        item = data['item']
                        if item.get('key') == key_text and has_comments(item):
                            # Add comment icon to Source Text
                            source_item = self.table.item(row, source_col)
                            if source_item and not source_item.text().startswith('💬 '):
                                source_item.setText('💬 ' + source_item.text().replace('💬 ', ''))
                            break

            # Show status message
            self.statusBar.showMessage(
                f"Document processed: Added {comment_updates} comments from {match_results['matches']} matched entries.",
                5000
            )

            # Show a popup to inform user
            QMessageBox.information(
                self,
                "Document Processing Complete",
                f"Successfully processed the document:\n\n"
                f"- Found {len(result.get('tables', []))} tables with conversation data\n"
                f"- Matched {match_results['matches']} entries with MXLIFF content\n"
                f"- Updated {comment_updates} entries with comments\n\n"
                f"Comments will be shown when hovering over entries or included in export."
            )
        else:
            # Error already handled in _on_worker_error
            pass

    def on_item_selected(self, current, previous):
        """Update tooltip for a selected item if it has comments."""
        if not current:
            return

        row = current.row()
        key_col = self.column_map.get('Key', 0)
        key_item = self.table.item(row, key_col)

        if not key_item:
            return

        key_text = key_item.text()

        # Find the corresponding data item
        data_item = None
        for data in self.processed_data:
            if not data.get('is_header', True) and 'item' in data:
                item = data['item']
                if item.get('key', '') == key_text:
                    data_item = item
                    break

        if data_item and has_comments(data_item):
            # Update tooltip for this row only
            tooltip_text = data_item.get('note_text', '')
            comment_text = get_comment_text(data_item)
            if comment_text:
                tooltip_text = f"{comment_text}\n\n{tooltip_text}"

            # Update cells in this row
            for col in range(self.table.columnCount()):
                cell = self.table.item(row, col)
                if cell:
                    cell.setToolTip(tooltip_text)



    def cancel_current_operation(self):
        """Cancel the current worker operation if any."""
        if hasattr(self, 'worker') and self.worker.isRunning():
            # Ask for confirmation
            reply = QMessageBox.question(
                self,
                "Cancel Operation",
                "Do you want to cancel the current operation?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )

            if reply == QMessageBox.Yes:
                self.worker.cancel()

                # Update UI to reflect cancellation
                self.progress_bar.setVisible(False)
                if hasattr(self, 'cancel_button'):
                    self.cancel_button.hide()

                self.statusBar.showMessage("Operation cancelled", 3000)


    def load_fonts(self):
        """Load custom fonts for the application."""
        # These fonts are typically available on most systems
        # For a commercial app, you would bundle custom fonts with your app
        self.title_font = QFont("Segoe UI", 16, QFont.Bold)
        self.header_font = QFont("Segoe UI", 11, QFont.Bold)
        self.normal_font = QFont("Segoe UI", 10)
        self.small_font = QFont("Segoe UI", 8)  # Added smaller font for word count display
        self.mono_font = QFont("Consolas", 9)

        # Create a fonts dictionary for easier access
        self.fonts = {
            'title': self.title_font,
            'header': self.header_font,
            'normal': self.normal_font,
            'small': self.small_font,
            'mono': self.mono_font
        }

    def initUI(self):
        # Set window properties
        self.setWindowTitle('D4 Scripts Tool')
        self.setGeometry(100, 100, 1280, 800)

        # Create UI components helper
        self.ui_components = UIComponents(self, self.fonts, self.current_columns)

        if hasattr(self, 'export_excel_button'):
            self.log("Connecting Excel export button in initUI")
            self.export_excel_button.clicked.disconnect()  # Disconnect any existing connections
            self.export_excel_button.clicked.connect(self.export_to_excel)

            # Create application icon
        app_icon = self.ui_components.create_app_icon()
        self.setWindowIcon(app_icon)

        # Create main layout structure
        central_widget = QWidget()
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(0)
        main_layout.setContentsMargins(0, 0, 0, 0)

        # Create toolbar
        self.create_toolbar()

        # Create header section
        header_widget = self.create_header_section()

        # Table panel
        table_widget = self.ui_components.create_table_panel()

        # Status bar
        self.statusBar = QStatusBar()
        self.statusBar.setFont(self.normal_font)
        self.setStatusBar(self.statusBar)
        self.statusBar.showMessage('Ready')

        # Add widgets to main layout
        main_layout.addWidget(header_widget)
        main_layout.addWidget(table_widget, 1)  # 1 = stretch factor

        # Set central widget
        self.setCentralWidget(central_widget)

        # Connect double click event to the table
        self.table.cellDoubleClicked.connect(self.on_table_double_clicked)

        # In your initUI method
        self.table.selectionModel().currentChanged.connect(self.on_item_selected)

        # Set up table columns
        self.setup_table_columns()

        # Apply theme
        self.apply_theme()

    def create_toolbar(self):
        """Create a toolbar with actions and resources dropdown."""
        self.toolbar = QToolBar("Main Toolbar")
        self.toolbar.setIconSize(QSize(16, 16))
        self.toolbar.setMovable(False)
        self.addToolBar(self.toolbar)

        # Open file action
        open_action = QAction("Open", self)
        open_action.triggered.connect(self.open_file)
        open_action.setStatusTip("Open an MXLIFF file")
        self.toolbar.addAction(open_action)

        # Export MXLIFF file action
        self.export_action = QAction("Export MXLIFF", self)
        self.export_action.triggered.connect(self.export_file)
        self.export_action.setStatusTip("Export the edited MXLIFF file")
        self.export_action.setEnabled(False)  # Disabled until a file is loaded
        self.toolbar.addAction(self.export_action)

        # NEW: Add Excel export action
        self.export_excel_action = QAction("Export to Excel", self)
        self.export_excel_action.triggered.connect(self.export_to_excel)
        self.export_excel_action.setStatusTip("Export the data to Excel with formatting")
        self.export_excel_action.setEnabled(False)  # Disabled until a file is loaded
        self.toolbar.addAction(self.export_excel_action)



        # Script Resources dropdown
        resources_menu = QMenu("Script Resources", self)

        # Content Team Info action
        content_team_action = QAction("Content Team Info", self)
        content_team_action.triggered.connect(self.open_content_team_info)
        resources_menu.addAction(content_team_action)

        # Queries action
        queries_action = QAction("Queries", self)
        queries_action.triggered.connect(self.open_queries)
        resources_menu.addAction(queries_action)

        # Resources dropdown button
        resources_button = QToolButton(self)
        resources_button.setText("Script Resources")
        resources_button.setMenu(resources_menu)
        resources_button.setPopupMode(QToolButton.InstantPopup)

        self.toolbar.addWidget(resources_button)

        # Add right-aligned spacer
        spacer = QWidget()
        spacer.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.toolbar.addWidget(spacer)

        # About action
        about_action = QAction("About", self)
        about_action.triggered.connect(self.show_about)
        self.toolbar.addAction(about_action)

    def create_header_section(self):
        """Create the application header section."""
        header_frame = QFrame()
        header_frame.setObjectName("headerFrame")
        header_frame.setFixedHeight(120)

        header_layout = QVBoxLayout(header_frame)
        header_layout.setContentsMargins(20, 10, 20, 10)
        header_layout.setSpacing(8)

        # Top row with title and version
        top_row = QHBoxLayout()

        # Title label
        title_label = QLabel('D4 Scripts Tool')
        title_label.setObjectName("titleLabel")
        title_label.setFont(self.title_font)
        top_row.addWidget(title_label)

        # Version label (right-aligned)
        version_label = QLabel("v1.7")  # Updated version to reflect the Excel export feature
        version_label.setObjectName("versionLabel")
        version_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        top_row.addWidget(version_label)

        header_layout.addLayout(top_row)

        # Middle section with file operations
        middle_row = QHBoxLayout()
        middle_row.setSpacing(10)

        # Buttons container
        buttons_container = QWidget()
        buttons_layout = QHBoxLayout(buttons_container)
        buttons_layout.setContentsMargins(0, 0, 0, 0)
        buttons_layout.setSpacing(10)

        # Open file button
        self.open_button = QPushButton('Open MXLIFF File')
        self.open_button.setObjectName("primaryButton")
        self.open_button.clicked.connect(self.open_file)
        self.open_button.setFixedSize(160, 40)  # Slightly smaller to fit all buttons
        buttons_layout.addWidget(self.open_button)

        # Export file button
        self.export_button = QPushButton('Export MXLIFF')
        self.export_button.setObjectName("primaryButton")
        self.export_button.clicked.connect(self.export_file)
        self.export_button.setFixedSize(160, 40)
        self.export_button.setEnabled(False)  # Disabled until a file is loaded
        buttons_layout.addWidget(self.export_button)

        self.export_excel_button = QPushButton('Export to Excel')
        self.export_excel_button.setObjectName("primaryButton")  # Ensure this matches other buttons
        self.export_excel_button.setCursor(Qt.PointingHandCursor)  # Add cursor change on hover
        self.export_excel_button.clicked.connect(self.export_to_excel)
        self.export_excel_button.setFixedSize(160, 40)
        self.export_excel_button.setEnabled(False)
        # Apply the same style explicitly to ensure hover works
        self.export_excel_button.setStyleSheet("")  # Clear any custom styles that might be interfering
        buttons_layout.addWidget(self.export_excel_button)

        # Upload Document button
        self.upload_doc_button = QPushButton('Upload Document')
        self.upload_doc_button.setObjectName("primaryButton")
        self.upload_doc_button.clicked.connect(self.upload_document)
        self.upload_doc_button.setFixedSize(160, 40)
        buttons_layout.addWidget(self.upload_doc_button)

        # Add buttons container to middle row
        middle_row.addWidget(buttons_container)
        middle_row.addStretch(1)

        header_layout.addLayout(middle_row)

        # Bottom row with file info and progress
        bottom_row = QHBoxLayout()

        # File name label
        self.file_label = QLabel('No file selected')
        self.file_label.setObjectName("fileLabel")
        self.file_label.setWordWrap(True)
        bottom_row.addWidget(self.file_label)

        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setObjectName("progressBar")
        self.progress_bar.setVisible(False)
        self.progress_bar.setRange(0, 0)  # Indeterminate progress
        self.progress_bar.setFixedWidth(200)
        bottom_row.addWidget(self.progress_bar)

        header_layout.addLayout(bottom_row)

        return header_frame

    def setup_table_columns(self):
        """Configure table columns based on current column order."""
        # Map of columns to their preferred width
        width_map = {
            'Key': 200,
            'Info': 40,
            'Speaker': 150,
            'Source Text': QHeaderView.Stretch,
            'Target Text': QHeaderView.Stretch,
            'Char Info': 150,  # Set width for the new column
            'Speaker and Target': 200,
            'Player Info': 200
        }

        # Apply column widths based on current order
        for i, col_name in enumerate(self.current_columns):
            if col_name in width_map:
                if width_map[col_name] == QHeaderView.Stretch:
                    self.header.setSectionResizeMode(i, QHeaderView.Stretch)
                else:
                    self.header.setSectionResizeMode(i, QHeaderView.Interactive)
                    self.table.setColumnWidth(i, width_map[col_name])

        # Update column map
        self.column_map = {col: idx for idx, col in enumerate(self.current_columns)}

        # Enable text wrapping for all columns
        self.table.setWordWrap(True)

        # Adjust row height to accommodate wrapped text
        self.table.resizeRowsToContents()

    def apply_theme(self):
        """Apply the current theme to all UI elements."""
        stylesheet = ThemeManager.generate_stylesheet(self.current_theme)
        self.setStyleSheet(stylesheet)
        self.update_table_colors()

    def update_table_colors(self):
        """Update the group header colors in the table based on the current theme."""
        if not hasattr(self, 'group_headers') or not self.group_headers:
            return

        # Update header colors
        for header_info in self.group_headers:
            row = header_info.get('row')
            if row is not None and row < self.table.rowCount():
                cell = self.table.item(row, 0)
                if cell:
                    # MODIFIED: Always use the group_header color regardless of MenuLabel status
                    cell.setBackground(QColor(self.current_theme['group_header']))

    def log(self, message):
        """Add a message to the log (now just prints to stdout)."""
        print(message)

    def has_unsaved_changes(self):
        """Check if there are unsaved changes in the data."""
        if not hasattr(self, 'processed_data') or not self.processed_data:
            return False

        # For each data item, check if the target text differs from the original
        for data in self.processed_data:
            if not data.get('is_header', True) and 'item' in data:
                item = data['item']
                key = item.get('key', '')

                # Get the item from the table to compare
                if key:
                    target_text = item.get('target_text', '')
                    original_text = item.get('original_target_text', target_text)

                    # If they differ, we have unsaved changes
                    if target_text != original_text:
                        return True

        return False

    def open_content_team_info(self):
        """Open Content Team Info link."""
        webbrowser.open(
            "https://confluence.blizzard.com/pages/viewpage.action?pageId=1586946082#CoTD42.3.0(Season9)-Updatedon:2025/3/31")

    def open_queries(self):
        """Open Queries link."""
        webbrowser.open(
            "https://blizzard.sharepoint.com/:x:/r/sites/GlobalLocalizationOps/_layouts/15/Doc.aspx?sourcedoc=%7B13B14CB9-2F4E-42E6-ACA6-060B4D58353C%7D&file=Diablo%204.xlsm&action=default&mobileredirect=true")

    def show_about(self):
        """Show the about dialog."""
        QMessageBox.about(
            self,
            "About D4 Scripts Tool",
            "<h2>D4 Scripts Tool</h2>"
            "<p>A tool for parsing and visualizing MXLIFF files.</p>"
            "<p>Built with PyQt5.</p>"
            "<p>&copy; 2025 Blizzard Localization</p>"
        )

    def open_file(self):
        """Open an MXLIFF file and parse it."""
        # Check for unsaved changes
        if self.has_unsaved_changes():
            reply = QMessageBox.question(
                self,
                "Unsaved Changes",
                "You have unsaved changes. Do you want to save them before opening a new file?",
                QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel,
                QMessageBox.Save
            )

            if reply == QMessageBox.Save:
                self.export_file()
                # If user cancels during export, abort opening
                if self.has_unsaved_changes():
                    return
            elif reply == QMessageBox.Cancel:  # Fixed: this should be at the same level as the if
                return

        # Fixed: This should be outside the if block, at the same level as the initial if statement
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Open MXLIFF File",
            "",
            "MXLIFF Files (*.mxliff *.xml);;All Files (*)",
            options=options
        )

        if file_path:
            self.log(f"Opening file: {file_path}")
            self.file_label.setText(os.path.basename(file_path))
            self.progress_bar.setVisible(True)
            self.statusBar.showMessage("Processing file...")

            # Clear previous results
            self.table.setRowCount(0)
            self.processed_data = []
            self.diff_pairs = {}

            # Parse the file in a separate timer to avoid freezing UI
            QTimer.singleShot(100, lambda: self._process_file(file_path))

    def _process_file(self, file_path):
        """Process the file after a short delay to allow UI to update."""
        try:
            # Parse the file
            self.log("Starting to parse the file...")

            with codecs.open(file_path, 'r', 'utf-8', errors='ignore') as f:
                xml_content = f.read()

                # Store the original content and file path
                self.original_xml_content = xml_content
                self.current_file_path = file_path

                # Direct XML parsing approach
                self.parse_xml(xml_content)

            self.progress_bar.setVisible(False)

            # Update table stats
            self.table_stats.setText(f"{self.table.rowCount()} entries")

            # Check for missing lines AFTER parsing
            missing_lines = self.check_missing_lines()

            # Show success message with stats
            success_message = f"File processed successfully. {self.table.rowCount()} entries loaded."
            if missing_lines:
                # If there are missing lines, modify the success message
                total_missing = sum(len(group['missing_lines']) for group in missing_lines)
                success_message += f" (Missing lines: {total_missing})"

            self.statusBar.showMessage(success_message, 5000)

            # Enable export controls now that we have a file loaded
            if hasattr(self, 'export_button'):
                self.export_button.setEnabled(True)
            if hasattr(self, 'export_action'):
                self.export_action.setEnabled(True)

        # Enable Excel export controls as well
            if hasattr(self, 'export_excel_button'):
                self.export_excel_button.setEnabled(True)
                # Optional: Add logging to confirm button is enabled
                self.log("Excel export button enabled")
            if hasattr(self, 'export_excel_action'):
                self.export_excel_action.setEnabled(True)

        except Exception as e:
            self.progress_bar.setVisible(False)
            error_msg = f"Error: {str(e)}"
            self.file_label.setText(error_msg)

            # Print full traceback to log
            self.log(f"Exception: {str(e)}")
            self.log(traceback.format_exc())

            # Update status bar
            self.statusBar.showMessage("Error processing file", 5000)

            # Show detailed error message
            QMessageBox.critical(
                self,
                "Error Processing File",
                f"{error_msg}\n\n"
                f"Check the debug log for more details."
            )

    def export_file(self):
        """Export the updated MXLIFF file with edited translations."""
        if not self.original_xml_content or not self.current_file_path:
            QMessageBox.warning(
                self,
                "Export Error",
                "No file is currently loaded or the original content is missing."
            )
            return

        # Check if we have any changes to export
        has_changes = False
        for data in self.processed_data:
            if not data.get('is_header', True) and 'item' in data:
                item = data['item']
                if 'target_text' in item and 'original_target_text' in item:
                    if item['target_text'] != item['original_target_text']:
                        has_changes = True
                        break

        if not has_changes:
            reply = QMessageBox.question(
                self,
                "No Changes Detected",
                "No changes to translations were detected. Do you still want to export the file?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply == QMessageBox.No:
                return

        # Choose where to save the file
        options = QFileDialog.Options()
        default_name = "edited_" + os.path.basename(self.current_file_path)
        save_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export to MXLIFF File",
            default_name,
            "MXLIFF Files (*.mxliff *.xml);;All Files (*)",
            options=options
        )

        if not save_path:
            return  # User cancelled

        try:
            # Show progress
            self.progress_bar.setVisible(True)
            self.statusBar.showMessage("Exporting file...")

            # We'll update the XML in a timer to allow UI to update
            QTimer.singleShot(100, lambda: self._process_export(save_path))

        except Exception as e:
            self.progress_bar.setVisible(False)
            error_msg = f"Error preparing export: {str(e)}"
            self.statusBar.showMessage(error_msg, 5000)

            QMessageBox.critical(
                self,
                "Export Error",
                error_msg
            )

    def _process_export(self, save_path):
        """Process the export after a short delay to allow UI to update."""
        try:
            # Show edited translations in log (for debugging)
            edited_count = 0
            for data in self.processed_data:
                if not data.get('is_header', True) and 'item' in data:
                    item = data['item']
                    if 'target_text' in item and 'original_target_text' in item:
                        if item['target_text'] != item['original_target_text']:
                            edited_count += 1
                            self.log(f"Edited: Key={item.get('key', 'UNKNOWN')}")
                            self.log(f"  Original: {item.get('original_target_text', '')}")
                            self.log(f"  Changed: {item.get('target_text', '')}")
                            if edited_count <= 5:  # Log only first 5 for brevity
                                self.log(
                                    f"Edited key: {item.get('key', '')} - Original: '{item.get('original_target_text', '')}' -> New: '{item.get('target_text', '')}'")

            self.log(f"Total edited translations: {edited_count}")

            # Create an updated XML content with the edited translations
            self.log("Starting XML update...")
            updated_xml = XMLParser.update_xml_content(self.original_xml_content, self.processed_data, self.log)
            self.log("XML update completed")

            # Write to the new file
            with codecs.open(save_path, 'w', 'utf-8') as f:
                f.write(updated_xml)

            self.progress_bar.setVisible(False)
            self.statusBar.showMessage(f"File exported successfully to {save_path}", 5000)

            # Provide a detailed success message
            QMessageBox.information(
                self,
                "Export Successful",
                f"MXLIFF file was successfully exported to:\n{save_path}\n\n"
                f"Total translations: {len(self.processed_data) - len(self.group_headers)}\n"
                f"Edited translations: {edited_count}"
            )

        except Exception as e:
            self.progress_bar.setVisible(False)
            error_msg = f"Error exporting file: {str(e)}"
            self.statusBar.showMessage(error_msg, 5000)

            # Print full traceback to log
            self.log(f"Export Exception: {str(e)}")
            self.log(traceback.format_exc())

            QMessageBox.critical(
                self,
                "Export Error",
                error_msg
            )

    def upload_document(self):
        """Handle uploading and processing a Word or PDF document."""
        # Check if we have MXLIFF data loaded
        if not hasattr(self, 'processed_data') or not self.processed_data:
            QMessageBox.warning(
                self,
                "No MXLIFF Data",
                "Please open an MXLIFF file first before uploading a document."
            )
            return

        # Initialize document parser
        self.document_parser = DocumentParser(self)

        # Select document file
        file_path = self.document_parser.select_document()
        if not file_path:
            return  # User cancelled

        # Show progress
        self.progress_bar.setVisible(True)
        self.statusBar.showMessage(f"Processing document: {os.path.basename(file_path)}...")

        # Process document in a separate timer to avoid UI freezing
        QTimer.singleShot(100, lambda: self._process_document())

    def _process_document(self):
        """
        Modified document processing method with enhanced debugging.
        Ensures Info column icons are updated when document matches are found.
        """
        try:
            # Parse the document
            success = self.document_parser.parse_document()
            if not success:
                self.progress_bar.setVisible(False)
                self.statusBar.showMessage("Failed to parse document.", 5000)
                return

            # Get tables with Conversation column
            conversation_tables = self.document_parser.get_conversation_tables()
            if not conversation_tables:
                self.progress_bar.setVisible(False)
                QMessageBox.warning(
                    self,
                    "No Conversation Tables Found",
                    "The document does not contain any tables with a 'Conversation' column."
                )
                self.statusBar.showMessage("No conversation tables found in document.", 5000)
                return

            # Match content with MXLIFF data
            match_results = self.document_parser.match_content_with_mxliff(self.processed_data)

            # Track updated keys
            updated_keys = set()
            comment_updates = 0

            # Debugging: Log all match results
            self.log("Match Results:")
            for update in match_results['updates']:
                self.log(f"Key: {update.get('key')}, Comment: {update.get('comment')}")

            # Update MXLIFF data with comments
            for update in match_results['updates']:
                key = update.get('key')
                comment = update.get('comment')

                # Debug: Try to match even without .F
                if key.endswith('.F'):
                    base_key = key[:-2]
                else:
                    base_key = key
                    key_variants = [key, f"{key}.F"]

                # Find matching keys (including variants)
                matched_keys = [key]
                if key.endswith('.F'):
                    matched_keys.append(base_key)
                elif not key.endswith('.F'):
                    matched_keys.append(f"{key}.F")

            # Update all matching keys
            for match_key in matched_keys:
                # Find the item in processed data
                for data in self.processed_data:
                    if not data.get('is_header', True) and 'item' in data:
                        item = data['item']
                        if item.get('key') == match_key:
                            # Add comment to note_text field
                            note_text = item.get('note_text', '')

                            # Check if there's already a comment
                            if 'Comment:' in note_text:
                                # Replace existing comment
                                note_text = re.sub(
                                    r'Comment:.*?(?=\n|$)',
                                    f'Comment: {comment}',
                                    note_text
                                )
                            else:
                                # Add new comment
                                if note_text:
                                    note_text += f"\nComment: {comment}"
                                else:
                                    note_text = f"Comment: {comment}"

                            # Update the note_text
                            item['note_text'] = note_text
                            item['has_document_match'] = True  # Add this line here
                            comment_updates += 1
                            updated_keys.add(match_key)
                            self.log(f"Updated key: {match_key} with comment")
                            break

            # Get the Info column index for verification
            info_col = self.column_map.get('Info', 1)
            if info_col >= 0:
                self.log(f"Info column found at index {info_col}, will update with icons")
            else:
                self.log("Warning: Info column not found in column map!")

            # Efficiently update comments and info icons in the table
            self.update_comments_efficiently(updated_keys)

            # Hide progress bar and show success message
            self.progress_bar.setVisible(False)

            # Show result message
            if comment_updates > 0:
                self.statusBar.showMessage(
                    f"Document processed: Added {comment_updates} comments from {match_results['matches']} matched entries.",
                    5000
                )
                QMessageBox.information(
                    self,
                    "Document Processing Complete",
                    f"Successfully processed the document:\n\n"
                    f"- Found {len(conversation_tables)} tables with conversation data\n"
                    f"- Matched {match_results['matches']} entries with MXLIFF content\n"
                    f"- Updated {comment_updates} entries with comments\n"
                    f"- Added info icons to {len(updated_keys)} matched entries"  # Added this line
                )
            else:
                self.statusBar.showMessage("Document processed: No matching entries found.", 5000)
                QMessageBox.information(
                    self,
                    "Document Processing Complete",
                    f"Processed the document, but found no matching entries to update."
                )

        except Exception as e:
            # Log and show error
            self.log(f"Document processing error: {str(e)}")
            self.log(traceback.format_exc())

            self.progress_bar.setVisible(False)
            self.statusBar.showMessage(f"Error processing document: {str(e)}", 5000)

            QMessageBox.critical(
                self,
                "Document Processing Error",
                f"An error occurred while processing the document:\n\n{str(e)}"
            )

    def update_comments_efficiently(self, updated_keys):
        """
        Efficiently update comments for specific keys without redrawing entire table.
        Now also adds info icons to the Info column when matching results are found.

        Args:
            updated_keys (set): Set of keys that have been updated with new comments
        """
        if not updated_keys:
            return

        # Columns we care about
        key_col = self.column_map.get('Key', 0)
        info_col = self.column_map.get('Info', 1)  # Get the Info column index
        source_col = self.column_map.get('Source Text', 2)  # Adjust index if needed due to new Info column

        # Iterate through table rows
        for row in range(self.table.rowCount()):
            key_item = self.table.item(row, key_col)
            if not key_item:
                continue

            key_text = key_item.text()

            # Check if this key is in the updated keys
            if key_text in updated_keys:
                # Find the corresponding data item in processed_data
                matching_data = next(
                    (data for data in self.processed_data
                     if not data.get('is_header', False)
                     and data.get('item', {}).get('key') == key_text),
                    None
                )

                if matching_data:
                    item = matching_data['item']

                    # Prepare tooltip with note text and comments
                    note_text = item.get('note_text', '')
                    comment_text = get_comment_text(item) if has_comments(item) else ''

                    tooltip_text = note_text
                    if comment_text:
                        tooltip_text = f"{comment_text}\n\n{note_text}"

                    # Update tooltip for all cells in this row
                    for col in range(self.table.columnCount()):
                        if col != info_col:  # Skip info column as we'll handle it separately
                            cell = self.table.item(row, col)
                            if cell:
                                cell.setToolTip(tooltip_text)

                    # Update the Info column with info icon
                    if info_col >= 0:
                        # Remove any existing cell widget
                        existing_widget = self.table.cellWidget(row, info_col)
                        if existing_widget:
                            self.table.removeCellWidget(row, info_col)

                        # Create and set the info icon with tooltip
                        self.create_info_icon(row, info_col, tooltip_text)

                        # Log that we've added an icon (for debugging)
                        self.log(f"Added info icon for key: {key_text}")

                    # Update the source text column with comment icon if needed
                    if source_col >= 0:
                        source_item = self.table.item(row, source_col)
                        if source_item:
                            current_text = source_item.text()
                            if has_comments(item) and not current_text.startswith('💬 '):
                                source_item.setText("💬 " + current_text.replace('💬 ', ''))
                            elif not has_comments(item) and current_text.startswith('💬 '):
                                # Remove icon if comment was removed
                                source_item.setText(current_text.replace('💬 ', ''))

    def check_missing_lines(self):
        """
        Check and log missing lines across all dialogue groups.
        Can be called after parsing to provide a comprehensive report.
        """
        # Dictionary to track missing lines for each group
        missing_lines_by_group = {}

        # Iterate through processed data
        for data in self.processed_data:
            # Skip header rows
            if data.get('is_header', False):
                continue

            # Get the item
            item = data.get('item', {})

            # Check if this is a missing line
            if item.get('is_missing_line', False):
                # Extract the main key (dialogue group)
                main_key = extract_main_key(item.get('key', ''))

                # If main key is empty, use a default
                if not main_key:
                    main_key = "UngroupedContent"

                # Initialize group entry if not exists
                if main_key not in missing_lines_by_group:
                    missing_lines_by_group[main_key] = []

                # Add the missing line number
                missing_lines_by_group[main_key].append(item.get('missing_line_number', '?'))

        # If no missing lines, return early
        if not missing_lines_by_group:
            return []

        # Prepare detailed missing lines report
        missing_lines_report = []

        # Prepare a user-friendly message
        message = "Missing Lines Detected:\n"
        for group, lines in missing_lines_by_group.items():
            # Sort missing line numbers
            sorted_lines = sorted(lines)

            # Create a report entry
            report_entry = {
                'group': group,
                'missing_lines': sorted_lines
            }
            missing_lines_report.append(report_entry)

            # Build message string
            message += f"- Group '{group}': Line(s) {', '.join(map(str, sorted_lines))}\n"

        # Show a warning dialog if there are missing lines
        if missing_lines_by_group:
            QMessageBox.warning(
                self,
                "Missing Dialogue Lines",
                message
            )

            # Update status bar
            self.statusBar.showMessage(
                f"Missing lines detected in {len(missing_lines_by_group)} groups",
                5000
            )

        return missing_lines_report

    def compare_translations(self, display_data):
        """
        Compares translations with matching keys (where one has .F suffix)
        and stores information about the differences for highlighting.
        """
        # Create a dictionary to store all items by their keys
        key_map = {}

        # First pass: Map all items by their keys
        for data in display_data:
            if not data['is_header'] and 'item' in data:
                item = data['item']
                key = item.get('key', '')
                if key:
                    key_map[key] = item

        # Second pass: Find pairs and store difference information
        self.diff_pairs = {}  # Store pairs of related translations

        for key in key_map:
            # Check if this key has a female variant
            if not key.endswith('.F'):
                female_key = f"{key}.F"
                if female_key in key_map:
                    # We found a pair!
                    male_item = key_map[key]
                    female_item = key_map[female_key]

                    # Store the pair
                    self.diff_pairs[key] = {
                        'male': male_item,
                        'female': female_item,
                        'diffs': find_text_differences(
                            male_item.get('target_text', ''),
                            female_item.get('target_text', '')
                        )
                    }

        self.log(f"Found {len(self.diff_pairs)} pairs of translations with gender variations")
        return display_data

    def highlight_differences_in_table(self):
        """Highlight differences between gender variants."""
        if not self.diff_highlighting_enabled or not self.diff_pairs:
            return

        target_col = self.column_map.get('Target Text', 3)
        key_col = self.column_map.get('Key', 0)

        # Process all rows in the table
        for row in range(self.table.rowCount()):
            key_item = self.table.item(row, key_col)
            if not key_item:
                continue

            key_text = key_item.text()

            # Only process female variants (.F endings)
            if not key_text.endswith('.F'):
                continue

            # Find the base key
            base_key = key_text[:-2]

            # Check if this is part of a diff pair
            if base_key not in self.diff_pairs:
                continue

            # Get the differences
            diffs = self.diff_pairs[base_key].get('diffs', [])

            # Get the target text cell
            target_item = self.table.item(row, target_col)
            if not target_item:
                continue

            # Apply highlighting
            if diffs:
                target_item.setForeground(QColor(self.current_theme.get('diff_text', '#FF0000')))
                target_item.setToolTip("Different words: " + ", ".join(diffs))
            else:
                target_item.setForeground(QColor(self.current_theme['text_primary']))
                target_item.setToolTip("")

    def find_row_by_key(self, key):
        """Find a row in the table by key."""
        key_col = self.column_map.get('Key', 0)
        for row in range(self.table.rowCount()):
            item = self.table.item(row, key_col)
            if item and item.text() == key:
                return row
        return -1

    def on_table_double_clicked(self, row, column):
        """Handle double click on a table cell."""
        try:
            # Check if the row and column are valid
            if row < 0 or column < 0 or row >= self.table.rowCount() or column >= self.table.columnCount():
                return

            # Get the item
            item = self.table.item(row, column)
            if not item:
                return

            # Check if it's a target text cell with differences
            if column == self.column_map.get('Target Text', 3):
                key_col = self.column_map.get('Key', 0)
                key_item = self.table.item(row, key_col)

                if key_item and key_item.text().endswith('.F'):
                    # Check if it has diff data
                    diff_words = item.data(Qt.UserRole + 1)
                    if diff_words:
                        self.show_diff_dialog(row)
        except Exception as e:
            # Log any errors
            print(f"Error in on_table_double_clicked: {str(e)}")

    def show_diff_dialog(self, row):
        """Show a dialog with highlighting of differences."""
        key_col = self.column_map.get('Key', 0)
        target_col = self.column_map.get('Target Text', 3)

        key_item = self.table.item(row, key_col)
        target_item = self.table.item(row, target_col)

        if not key_item or not target_item:
            return

        key_text = key_item.text()
        if not key_text.endswith('.F'):
            return

        # Get diff data
        diff_words = target_item.data(Qt.UserRole + 1)
        if not diff_words:
            return

        # Get the related base row
        base_row = target_item.data(Qt.UserRole + 2)
        if base_row < 0:
            return

        # Get base text
        base_target_item = self.table.item(base_row, target_col)
        if not base_target_item:
            return

        # Show dialog
        dialog = TranslationDiffDialog(
            self,
            base_target_item.text(),
            target_item.text(),
            diff_words
        )
        dialog.exec_()

    def on_selection_changed(self, selected, deselected):
        """Handle selection changes."""
        # No longer need to do anything here
        pass

    def on_cell_changed(self, row, column):
        """Handle changes to cell data."""
        # Prevent recursive calls
        if self.updating_cell:
            return

        # Set flag to prevent recursion
        self.updating_cell = True

        try:
            # Make sure it's the Target Text column
            target_col = self.column_map.get('Target Text', 3)
            if column != target_col:
                self.updating_cell = False
                return

            # Get the item that changed
            item = self.table.item(row, column)
            if not item:
                self.updating_cell = False
                return

            # Get the new text
            new_text = item.text()

            # Get the key for this row
            key_col = self.column_map.get('Key', 0)
            key_item = self.table.item(row, key_col)
            if not key_item:
                self.updating_cell = False
                return

            key_text = key_item.text()

            # Get original text
            original_text = item.data(Qt.UserRole)
            if original_text is None:
                self.updating_cell = False
                return

            # If nothing changed, do nothing
            if new_text == original_text:
                self.updating_cell = False
                return

            # Update our data structure
            updated = False
            has_comment = False

            for data in self.processed_data:
                if not data.get('is_header', True) and 'item' in data:
                    item_data = data['item']
                    if item_data.get('key', '') == key_text:
                        # Check if it has comments before we change it
                        has_comment = has_comments(item_data)

                        # Store original if not already stored
                        if 'original_target_text' not in item_data:
                            item_data['original_target_text'] = item_data.get('target_text', '')

                        # Update the text
                        item_data['target_text'] = new_text
                        updated = True
                        break

            if updated:
                # Store the new text for comparison later
                item.setData(Qt.UserRole, new_text)

                # Make text bold to show it's edited
                font = QFont(self.normal_font)
                font.setBold(True)
                item.setFont(font)

                # Update the Char Info column
                char_info_col = self.column_map.get('Char Info', 4)
                if char_info_col >= 0:
                    # Get source text
                    source_col = self.column_map.get('Source Text', 2)
                    source_item = self.table.item(row, source_col)
                    if not source_item:
                        return

                    source_text = source_item.text()

                    # Calculate character counts
                    source_chars = len(source_text)
                    target_chars = len(new_text)

                    # Calculate percentage difference
                    percentage = 0
                    if source_chars > 0:
                        percentage = ((target_chars - source_chars) / source_chars) * 100

                    # Create char info text
                    if source_chars == target_chars:
                        char_info = "Equal"
                    else:
                        if percentage > 0:
                            char_info = f"+{int(percentage)}%"
                        else:
                            char_info = f"{int(percentage)}%"

                    # Update char info cell
                    char_item = self.table.item(row, char_info_col)
                    if char_item:
                        char_item.setText(char_info)

                        # Color code
                        if abs(percentage) > 20:
                            char_item.setForeground(QColor(255, 0, 0))  # Red for big changes
                        elif percentage > 0:
                            char_item.setForeground(QColor(0, 128, 0))  # Green for positive expansions
                        elif percentage < 0:
                            char_item.setForeground(QColor(0, 0, 255))  # Blue for contractions
                        else:
                            char_item.setForeground(QColor(self.current_theme['text_primary']))

                    # Update the Source Text column with comment icon if needed
                    source_col = self.column_map.get('Source Text', 2)
                    if source_col >= 0 and has_comment:
                        source_item = self.table.item(row, source_col)
                        if source_item:
                            current_text = source_item.text()
                            if not current_text.startswith('💬 '):
                                source_item.setText("💬 " + current_text.replace('💬 ', ''))

                # Check if this is a female variant (.F)
                if key_text.endswith('.F'):
                    # Get the base key (without .F)
                    base_key = key_text[:-2]

                    # Look for the corresponding male variant row
                    male_row = -1
                    for r in range(self.table.rowCount()):
                        k_item = self.table.item(r, key_col)
                        if k_item and k_item.text() == base_key:
                            male_row = r
                            break

                    if male_row >= 0:
                        # Found the male variant row
                        male_item = self.table.item(male_row, target_col)
                        if male_item:
                            male_text = male_item.text()

                            # Compare texts
                            diffs = find_text_differences(male_text, new_text)

                            # Update diffs in the diff_pairs if it exists
                            if base_key in self.diff_pairs:
                                self.diff_pairs[base_key]['female']['target_text'] = new_text
                                self.diff_pairs[base_key]['diffs'] = diffs

                            # If they're different, highlight the female variant (this row)
                            if diffs:
                                item.setForeground(QColor(self.current_theme.get('diff_text', '#FF0000')))
                                item.setToolTip("Different words: " + ", ".join(diffs))
                            else:
                                item.setForeground(QColor(self.current_theme['text_primary']))
                                item.setToolTip("")

                else:
                    # This is a male variant - look for its female counterpart
                    female_key = f"{key_text}.F"

                    # Find the female variant row
                    female_row = -1
                    for r in range(self.table.rowCount()):
                        k_item = self.table.item(r, key_col)
                        if k_item and k_item.text() == female_key:
                            female_row = r
                            break

                    if female_row >= 0:
                        # Found the female variant row
                        female_item = self.table.item(female_row, target_col)
                        if female_item:
                            female_text = female_item.text()

                            # Compare texts
                            diffs = find_text_differences(new_text, female_text)

                            # Update diffs in the diff_pairs if it exists
                            if key_text in self.diff_pairs:
                                self.diff_pairs[key_text]['male']['target_text'] = new_text
                                self.diff_pairs[key_text]['diffs'] = diffs

                            # If they're different, highlight the female variant
                            if diffs:
                                female_item.setForeground(QColor(self.current_theme.get('diff_text', '#FF0000')))
                                female_item.setToolTip("Different words: " + ", ".join(diffs))
                            else:
                                female_item.setForeground(QColor(self.current_theme['text_primary']))
                                female_item.setToolTip("")

                # Status message
                self.statusBar.showMessage(f"Updated translation for '{key_text}'", 3000)

        except Exception as e:
            # Log any errors
            self.log(f"Error in cell changed: {str(e)}")
            self.log(traceback.format_exc())
        finally:
            # Always reset the flag
            self.updating_cell = False

    def update_char_count_info(self, row, column, text=None):
        """
        Calculate and display character count information.

        Args:
            row (int): Row index in the table
            column (int): Column index in the table
            text (str, optional): Text to analyze. If None, will extract from the table item.
        """
        # Ensure we're working with the Target Text column
        target_col = self.column_map.get('Target Text', 3)
        source_col = self.column_map.get('Source Text', 2)

        # Safety checks
        if (row < 0 or row >= self.table.rowCount() or
                column != target_col):
            return

        # Get table items
        target_item = self.table.item(row, target_col)
        source_item = self.table.item(row, source_col)

        if not target_item or not source_item:
            return

        # If no text provided, extract from the item
        if text is None:
            text = target_item.text()

        # Get source text
        source_text = source_item.text().strip()

        # Remove comment icon from source text if present
        if source_text.startswith('💬 '):
            source_text = source_text[2:].strip()

        # Count characters
        source_char_count = len(source_text)
        target_char_count = len(text)

        # Calculate percentage difference
        percentage_diff = 0
        if source_char_count > 0:
            percentage_diff = ((target_char_count - source_char_count) / source_char_count) * 100

        # Generate expansion text
        if abs(percentage_diff) < 1.0:
            expansion_text = "Same length as enUS"
        elif percentage_diff > 0:
            expansion_text = f"{int(percentage_diff + 0.5)}% longer than enUS"
        else:
            expansion_text = f"{int(abs(percentage_diff) + 0.5)}% shorter than enUS"

        # Prepare info text
        info_text = (
            f"Chars: {target_char_count} | "
            f"enUS: {source_char_count} | "
            f"Expansion: {expansion_text}"
        )

        # Store data in the item
        target_item.setData(Qt.UserRole + 3, info_text)

        # Color for significant changes
        color = None
        if abs(percentage_diff) > 20:
            color = QColor(255, 100, 100)  # Light red for big changes
        target_item.setData(Qt.UserRole + 4, color)

        # Force repaint to ensure visibility
        self.table.viewport().update()

        return info_text

    def parse_xml(self, xml_content):
        """Parse XML directly using regex approach for MXLIFF files."""
        self.log("Parsing XML using direct regex approach...")

        # Use the parser utility to process the XML content
        processed_data = XMLParser.parse_xml(xml_content, self.log)

        # Group by main key
        grouped_data = {}
        for item in processed_data:
            main_key = extract_main_key(item['key'])
            if not main_key:
                main_key = "UngroupedContent"
            if main_key not in grouped_data:
                grouped_data[main_key] = []
            grouped_data[main_key].append(item)

        self.log(f"Number of groups: {len(grouped_data)}")

        # Prepare data for display
        display_data = []

        # Sort keys using the custom natural sort key method
        sorted_keys = sorted(
            grouped_data.keys(),
            key=natural_sort_key
        )

        for main_key in sorted_keys:
            # Sort items within group by order value, then by original index
            group_items = sorted(
                grouped_data[main_key],
                key=lambda x: (x['order_value'], x['index'])
            )

            # Track the expected line order
            expected_line_number = 1
            ordered_group_items = []

            for item in group_items:
                # Extract the line number from the item
                current_line_number = extract_line_number(item)

                # Check for missing lines
                while expected_line_number < current_line_number:
                    # Create a placeholder for missing line
                    missing_line_item = {
                        **item,  # Copy most properties from the current item
                        'is_missing_line': True,
                        'missing_line_number': expected_line_number,
                        'source_text': '[MISSING LINE]',
                        'target_text': '[MISSING LINE]',
                        'key': f"{item.get('key', '')}_missing_{expected_line_number}"
                    }
                    ordered_group_items.append(missing_line_item)
                    expected_line_number += 1

                ordered_group_items.append(item)
                expected_line_number = current_line_number + 1

            # Add group header
            display_data.append({
                'is_header': True,
                'main_key': main_key,
                'item_count': len(ordered_group_items),
                'is_menulabel': any(item.get('is_menulabel', False) for item in ordered_group_items),
                'contains_menulabel': any('MenuLabel' in item.get('key', '') for item in ordered_group_items)
            })

            # Add items
            for item in ordered_group_items:
                display_data.append({
                    'is_header': False,
                    'item': item
                })

        # Store the processed data for reuse when column order changes
        self.processed_data = display_data

        # Display in table
        self.display_results(display_data)

    def display_results(self, display_data):
        """Display the parsed data in the table."""
        # First compare translations to find differences
        display_data = self.compare_translations(display_data)

        self.log(f"Displaying {len(display_data)} items")

        # DEBUG: Count headers vs content items
        headers = sum(1 for d in display_data if d.get('is_header', False))
        content_items = len(display_data) - headers
        self.log(f"Headers: {headers}, Content items: {content_items}")

        # Disconnect cell changed signal to prevent triggering during table setup
        try:
            self.table.cellChanged.disconnect(self.on_cell_changed)
        except:
            # It might not be connected yet
            pass

        # Clear previous data
        self.table.clearContents()
        self.group_headers = []
        self.group_rows = {}

        # Make sure we have data to display
        if not display_data:
            self.file_label.setText("No data found in the file. Please check if it's a valid MXLIFF file.")
            self.statusBar.showMessage("No data found", 5000)
            # Reconnect the signal before returning
            try:
                self.table.cellChanged.connect(self.on_cell_changed)
            except:
                pass
            return

        # Count how many scene info rows we will need by pre-scanning the data
        scene_info_count = 0
        for i, data in enumerate(display_data):
            if data.get('is_header', False) and i + 1 < len(display_data) and not display_data[i + 1].get('is_header',
                                                                                                          False):
                first_item = display_data[i + 1].get('item', {})
                note_text = first_item.get('note_text', '')
                if note_text and 'Scene:' in note_text:
                    scene_info_count += 1

        # Add extra rows for missing lines
        missing_line_count = sum(1 for data in display_data if not data.get('is_header', False) and
                                 data.get('item', {}).get('is_missing_line', False))

        # Set the table row count with extra space for scene info and missing lines
        total_rows = len(display_data) + scene_info_count + missing_line_count
        self.table.setRowCount(total_rows)
        self.log(
            f"Table row count set to: {total_rows} (data: {len(display_data)}, scene info: {scene_info_count}, missing lines: {missing_line_count})")

        # Track our progress through the table
        row_index = 0

        # Process each data item
        for data_index, data in enumerate(display_data):
            # Ensure we don't go beyond the allocated rows
            if row_index >= self.table.rowCount():
                # Dynamically resize the table if needed
                new_size = row_index + 10  # Add some buffer
                self.log(
                    f"WARNING: Row index {row_index} exceeds table rows {self.table.rowCount()}, resizing to {new_size}")
                self.table.setRowCount(new_size)

            if data.get('is_header', False):
                # Log for debugging
                self.log(f"Adding header at row {row_index}: {data.get('main_key', 'unknown')}")

                # Add a visible group header row to the table
                header_text = f"Dialogue Group: {data['main_key']} ({data['item_count']} entries)"
                header_cell = QTableWidgetItem(header_text)
                header_cell.setFlags(Qt.ItemIsEnabled)  # Not editable
                header_cell.setTextAlignment(Qt.AlignCenter)  # Center the text

                # Apply header styles
                header_cell.setBackground(QColor(self.current_theme['group_header']))
                header_cell.setFont(self.header_font)  # This font is bold

                # Store the group info in the cell data
                header_cell.setData(Qt.UserRole, {
                    'is_header': True,
                    'group': data['main_key'],
                    'is_menulabel': data.get('is_menulabel', False),
                    'contains_menulabel': data.get('contains_menulabel', False)
                })

                # Set the header cell in the first column
                self.table.setItem(row_index, 0, header_cell)

                # Make header span all columns
                try:
                    self.table.setSpan(row_index, 0, 1, self.table.columnCount())
                except Exception as e:
                    self.log(f"Error setting span: {str(e)}")

                # Store header row for later use
                self.group_headers.append({
                    'row': row_index,
                    'group': data['main_key'],
                    'header_cell': header_cell,
                    'is_menulabel': data.get('is_menulabel', False)
                })

                # IMPORTANT: Force a minimum row height for header
                self.table.setRowHeight(row_index, 30)

                row_index += 1

                # Check if there's scene info in the first item of this group
                scene_info = ""
                if data_index + 1 < len(display_data) and not display_data[data_index + 1].get('is_header', False):
                    first_item = display_data[data_index + 1].get('item', {})
                    note_text = first_item.get('note_text', '')
                    scene_match = re.search(r'Scene:\s*([^\n]+)', note_text)
                    if scene_match:
                        scene_info = f"Scene: {scene_match.group(1).strip()}"

                # Add scene info in a separate row if available
                if scene_info:
                    # Check row bounds again
                    if row_index >= self.table.rowCount():
                        new_size = row_index + 10
                        self.log(
                            f"WARNING: Row index {row_index} exceeds table rows {self.table.rowCount()}, resizing to {new_size}")
                        self.table.setRowCount(new_size)

                    # Create a new row for scene info
                    scene_cell = QTableWidgetItem(scene_info)
                    scene_cell.setFlags(Qt.ItemIsEnabled)  # Not editable
                    scene_cell.setTextAlignment(Qt.AlignCenter)  # Center the text

                    # Use normal font (not bold)
                    scene_cell.setFont(self.normal_font)

                    # Use a slightly lighter background to distinguish from header
                    lighter_color = QColor(self.current_theme['group_header'])
                    lighter_color.setAlpha(200)  # Make slightly transparent
                    scene_cell.setBackground(lighter_color)

                    # Insert the scene cell
                    self.table.setItem(row_index, 0, scene_cell)

                    # Make scene info span all columns
                    try:
                        self.table.setSpan(row_index, 0, 1, self.table.columnCount())
                    except Exception as e:
                        self.log(f"Error setting span: {str(e)}")

                    # IMPORTANT: Force a minimum row height for scene info
                    self.table.setRowHeight(row_index, 30)

                    row_index += 1

            else:
                # Check row bounds again
                if row_index >= self.table.rowCount():
                    new_size = row_index + 10
                    self.log(
                        f"WARNING: Row index {row_index} exceeds table rows {self.table.rowCount()}, resizing to {new_size}")
                    self.table.setRowCount(new_size)

                # Regular data row or missing line
                item = data['item']

                # Check if this is a missing line
                if item.get('is_missing_line', False):
                    # Create a special missing line row
                    missing_line_cell = QTableWidgetItem(f"[MISSING LINE {item.get('missing_line_number', '?')}]")
                    missing_line_cell.setFlags(Qt.ItemIsEnabled)  # Not editable
                    missing_line_cell.setBackground(QColor(255, 200, 200))  # Light red background
                    missing_line_cell.setForeground(QColor(150, 0, 0))  # Dark red text
                    missing_line_cell.setTextAlignment(Qt.AlignCenter)
                    missing_line_cell.setFont(self.header_font)  # Make it stand out

                    # Set the missing line cell in the first column
                    self.table.setItem(row_index, 0, missing_line_cell)

                    # Make missing line span all columns
                    try:
                        self.table.setSpan(row_index, 0, 1, self.table.columnCount())
                    except Exception as e:
                        self.log(f"Error setting span for missing line: {str(e)}")

                    row_index += 1
                    continue

                # Prepare data for each column based on current column layout
                # Combine Speaker Target and Speaker Gender
                speaker_info = []

                # Check for female key ending
                is_female_key = item.get('key', '') and str(item.get('key', '')).endswith('.F')

                # Add Speaker Target
                speaker_target = item.get('speaker_target', '')
                if speaker_target:
                    speaker_info.append(f"Speaker Target: {speaker_target}")
                elif is_female_key and not speaker_target:
                    speaker_info.append("Speaker Target: Player - Female")
                else:
                    speaker_info.append("Speaker Target: None")

                # Add Speaker Gender
                speaker_gender = item.get('speaker_gender', '')
                if speaker_gender and speaker_gender.lower() != 'none':
                    speaker_info.append(f"Speaker Gender: {speaker_gender}")
                else:
                    speaker_info.append("Speaker Gender: None")

                # Combine Player Class and Player Gender
                player_info = []

                # Add Player Class
                player_class = item.get('player_class', '')
                if player_class and player_class.lower() not in ('- none -', 'none', '-none-'):
                    player_info.append(f"Class: {player_class}")
                else:
                    player_info.append("Class: - None -")

                # Add Player Gender
                player_gender = item.get('player_gender', '')

                if player_gender and player_gender.lower() != 'none':
                    player_info.append(f"Gender: {player_gender}")
                elif is_female_key:
                    player_info.append("Gender: Female")
                else:
                    player_info.append("Gender: None")

                # Check if this item has comments that should be highlighted
                has_comment = has_comments(item)
                comment_text = get_comment_text(item) if has_comment else ""

                # Build source text with comment icon if needed
                source_text = item.get('source_text', '')
                if has_comment and not source_text.startswith('💬 '):
                    source_text = f"💬 {source_text}"

                # Prepare tooltip text
                tooltip_text = item.get('note_text', '')
                if has_comment:
                    tooltip_text = f"{comment_text}\n\n{tooltip_text}" if tooltip_text else comment_text

                # Inside the display_results method, modify the part that creates row data:
                row_data = {
                    'Key': item.get('key', ''),
                    'Info': '',  # Placeholder for Info column, we'll handle it separately
                    'Speaker': item.get('speaker', ''),
                    'Source Text': source_text,
                    'Target Text': item.get('target_text', ''),
                    'Char Info': '',  # We'll set this separately
                    'Speaker and Target': '\n'.join(speaker_info),
                    'Player Info': '\n'.join(player_info)
                }

                # Add data to table according to current column order
                for col_name, col_index in self.column_map.items():
                    if col_name == 'Char Info':
                        # Calculate character info for the Char Info column
                        target_text = item.get('target_text', '')
                        source_text = item.get('source_text', '')

                        source_char_count = len(source_text)
                        target_char_count = len(target_text)

                        # Calculate percentage difference
                        percentage_diff = 0
                        if source_char_count > 0:
                            percentage_diff = ((target_char_count - source_char_count) / source_char_count) * 100

                        # Prepare char info text
                        if source_char_count == target_char_count:
                            char_info = "Equal"
                        else:
                            if percentage_diff > 0:
                                char_info = f"+{int(percentage_diff)}%"
                            else:
                                char_info = f"{int(percentage_diff)}%"

                        # Create table item with char info
                        table_item = QTableWidgetItem(char_info)

                        # Make it non-editable
                        table_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)

                        # Color code if needed
                        if abs(percentage_diff) > 20:
                            table_item.setForeground(QColor(255, 0, 0))
                        elif percentage_diff > 0:
                            table_item.setForeground(QColor(0, 128, 0))  # Green for positive expansions
                        elif percentage_diff < 0:
                            table_item.setForeground(QColor(0, 0, 255))  # Blue for contractions

                    elif col_name == 'Info':
                        # Only show info icons when a document has been uploaded and matches found
                        if item.get('has_document_match', False):
                            # Create info icon with tooltip
                            self.create_info_icon(row_index, col_index, tooltip_text)
                            continue
                        else:
                            # Create empty cell
                            table_item = QTableWidgetItem("")
                            table_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)

                    else:
                        # Regular column processing
                        value = row_data.get(col_name, '')
                        table_item = QTableWidgetItem(value)

                        if col_name == 'Target Text':
                            # Make Target Text column editable
                            table_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
                            table_item.setData(Qt.UserRole, item.get('target_text', ''))
                        else:
                            # Other columns remain non-editable
                            table_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)

                    # Set the item in the table
                    self.table.setItem(row_index, col_index, table_item)
                    table_item.setFont(self.normal_font)

                # Set background color based on conditions
                bg_color = None

                # Check for MenuLabel in key
                if item.get('is_menulabel', False) or 'MenuLabel' in str(item.get('key', '')):
                    bg_color = QColor(self.current_theme['menu_label'])
                # Check for female key ending
                elif item.get('key', '') and str(item.get('key', '')).endswith('.F'):
                    bg_color = QColor(self.current_theme['female_key'])

                # Apply background color if needed
                if bg_color:
                    for col in range(self.table.columnCount()):
                        if self.table.item(row_index, col):
                            self.table.item(row_index, col).setBackground(bg_color)

                # Add tooltip with full note text and highlight comment if present
                if tooltip_text:
                    for col in range(self.table.columnCount()):
                        if col != self.column_map.get('Info', 1):  # Skip Info column as it already has tooltip
                            if self.table.item(row_index, col):
                                self.table.item(row_index, col).setToolTip(tooltip_text)

                # IMPORTANT: Force a minimum row height for data row
                self.table.setRowHeight(row_index, 30)

                row_index += 1

        # Apply difference highlighting
        self.highlight_differences_in_table()

        # Update table stats label
        self.table_stats.setText(f"{row_index} entries")

        # Force update to make sure character counts show up
        self.table.viewport().update()

        # Reconnect the signal after table setup is complete
        try:
            self.table.cellChanged.connect(self.on_cell_changed)
        except:
            pass

    # Override the paintEvent for QTableWidget to show word count info
    def eventFilter(self, obj, event):
        """Event filter to handle custom painting of table cells."""
        if obj is self.table.viewport() and event.type() == event.Paint:
            painter = QPainter(self.table.viewport())
            painter.setClipRect(event.rect())

            # Get visible rows range
            first_row = self.table.rowAt(0)
            last_row = self.table.rowAt(self.table.viewport().height())

            # Safety checks
            if first_row < 0:
                first_row = 0
            if last_row < 0:
                last_row = self.table.rowCount() - 1

            target_col = self.column_map.get('Target Text', 3)

            # Paint word count info below each target text cell for ALL editable cells
            for row in range(first_row, last_row + 1):
                item = self.table.item(row, target_col)

                # Check if the item exists and is editable
                if item and item.flags() & Qt.ItemIsEditable:
                    # Retrieve character count info
                    info_text = item.data(Qt.UserRole + 3)

                    # If no info text, try to calculate
                    if not info_text:
                        try:
                            # Force recalculation
                            self.update_char_count_info(row, target_col)
                            info_text = item.data(Qt.UserRole + 3)
                        except Exception as e:
                            print(f"Character count recalc error: {e}")
                            continue

                    if info_text:
                        rect = self.table.visualRect(self.table.model().index(row, target_col))

                        # Adjust position to bottom of cell
                        rect.setTop(rect.bottom() - 15)
                        rect.setLeft(rect.left() + 5)  # Add a small left margin

                        # Set color
                        color = item.data(Qt.UserRole + 4) or QColor(self.current_theme['word_count_text'])

                        painter.setPen(color)
                        painter.setFont(self.small_font)  # Use the small font defined in __init__

                        # Draw text with anti-aliasing for better readability
                        painter.setRenderHint(QPainter.TextAntialiasing)
                        painter.drawText(rect, Qt.AlignLeft | Qt.AlignBottom, info_text)

            painter.end()
            return False

        return super().eventFilter(obj, event)
